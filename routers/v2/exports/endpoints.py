import asyncio
import coc
import os
import tempfile
from fastapi import HTTPException, APIRouter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pendulum as pend

from routers.v2.war.models import PlayerWarhitsFilter
from routers.v2.war.utils import collect_player_hits_from_wars
from routers.v2.exports.utils import (
    insert_logo_from_cdn,
    format_table,
    build_cwl_info_data,
    build_cwl_member_row,
    generate_cwl_filename,
    generate_player_war_filename,
    build_filter_summary,
    extract_war_hits_from_results,
    export_player_war_stats_to_excel,
    CLASHKING_LOGO_URL
)
from utils.utils import fix_tag
from utils.database import MongoClient
from fastapi.responses import FileResponse

router = APIRouter(prefix="/v2/exports", tags=["Exports"], include_in_schema=True)

# Constants
CWL_MEMBER_HEADERS = [
    "Player Name", "Player Tag", "Town Hall", "Total Attacks", "Total Stars",
    "Average Stars", "Total Destruction %", "Average Destruction %", "Performance Score"
]
PREPARATION_START_TIME_FIELD = "data.preparationStartTime"


@router.get("/war/cwl-summary", name="Export CWL summary and members stats to Excel")
async def export_cwl_summary_to_excel(tag: str):
    try:
        clan_tag = fix_tag(tag)

        # Get current CWL season data
        pipeline = [
            {"$match": {"clan_tag": clan_tag}},
            {"$sort": {"season": -1}},
            {"$limit": 1}
        ]

        cursor = await MongoClient.clan_leaderboard.aggregate(pipeline)
        result = await cursor.to_list(length=1)
        if not result:
            raise HTTPException(status_code=404, detail="No CWL data found for this clan")

        cwl_data = result[0]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "CWL Summary"

        # Add logo
        await insert_logo_from_cdn(ws, CLASHKING_LOGO_URL, "J1", 50)

        # Title
        ws["A1"] = f"CWL Summary for {cwl_data.get('clan_name', 'Unknown')} - Season {cwl_data.get('season', 'Unknown')}"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:I1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Clan info summary using helper
        row = 3
        ws[f"A{row}"] = "Clan Information"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        info_data = build_cwl_info_data(cwl_data)
        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        # Add some spacing
        row += 2

        # Members performance table
        ws[f"A{row}"] = "Member Performance"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        # Headers
        for i, header in enumerate(CWL_MEMBER_HEADERS, 1):
            ws.cell(row=row, column=i, value=header)

        start_row = row

        # Add member data using helper
        for member in cwl_data.get('members', []):
            row += 1
            member_data = build_cwl_member_row(member)
            for i, value in enumerate(member_data, 1):
                ws.cell(row=row, column=i, value=value)

        end_row = row
        format_table(ws, start_row, end_row)

        # Save to temporary file (async-safe and secure approach)
        # Use mkstemp to create a secure temporary file
        fd, tmp_path = await asyncio.to_thread(
            tempfile.mkstemp,
            suffix=".xlsx",
            prefix="cwl_export_",
            dir=None,
            text=False
        )
        await asyncio.to_thread(wb.save, tmp_path)
        # Close the file descriptor as openpyxl handles the file
        await asyncio.to_thread(os.close, fd)

        # Generate filename using helper
        filename = generate_cwl_filename(
            cwl_data.get('clan_name', 'Clan'),
            clan_tag,
            cwl_data.get('season', 'unknown')
        )

        return FileResponse(
            path=tmp_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating CWL export: {str(e)}")


@router.post("/war/player-stats", name="Export player war statistics to Excel")
async def export_player_war_stats(filter: PlayerWarhitsFilter):
    # Get war hits data using existing logic
    client = coc.Client(raw_attribute=True)
    start = pend.from_timestamp(filter.timestamp_start, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')
    end = pend.from_timestamp(filter.timestamp_end, tz=pend.UTC).strftime('%Y%m%dT%H%M%S.000Z')

    player_tag = fix_tag(filter.player_tags[0])
    pipeline = [
        {"$match": {
            "$and": [
                {"$or": [
                    {"data.clan.members.tag": player_tag},
                    {"data.opponent.members.tag": player_tag}
                ]},
                {PREPARATION_START_TIME_FIELD: {"$gte": start}},
                {PREPARATION_START_TIME_FIELD: {"$lte": end}}
            ]
        }},
        {"$sort": {PREPARATION_START_TIME_FIELD: -1}},
        {"$limit": filter.limit or 10000},
        {"$unset": ["_id"]},
        {"$project": {"data": "$data"}},
    ]

    cursor = await MongoClient.clan_wars.aggregate(pipeline, allowDiskUse=True)
    wars_docs = await cursor.to_list(length=None)

    result = await collect_player_hits_from_wars(
        wars_docs,
        tags_to_include=[player_tag],
        clan_tags=None,
        filter=filter,
        client=client
    )

    player_results = result["items"]

    if not player_results:
        raise HTTPException(status_code=404, detail="No war hits found for the specified player and filters")

    # Extract individual attacks from war data using helper
    war_hits = extract_war_hits_from_results(player_results)

    if not war_hits:
        raise HTTPException(status_code=404, detail="No individual attacks found for the specified player and filters")

    # Build filter summary using helper
    filter_summary = build_filter_summary(filter)

    # Use the new modular export function
    tmp = await export_player_war_stats_to_excel(
        war_hits=war_hits,
        player_tag=player_tag,
        filter_summary=filter_summary
    )

    # Generate filename and return file
    player_name = next((hit.get("attacker_name", "") for hit in war_hits if hit.get("attacker_name")), "Player")
    filename = generate_player_war_filename(player_name, player_tag)

    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )