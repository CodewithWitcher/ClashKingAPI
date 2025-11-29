"""
Excel export utilities for ClashKing API
"""
import aiohttp
import tempfile
from tempfile import NamedTemporaryFile
from typing import Optional, List, Dict, Any, Tuple
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

# ClashKing theme colors and styles
CLASHKING_THEME = {
    "header_fill_red": PatternFill(start_color="D00000", end_color="D00000", fill_type="solid"),
    "header_fill_black": PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
    "header_font_white": Font(color="FFFFFF", bold=True),
    "data_font_black": Font(color="000000"),
    "title_fill": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
    "data_fill_white": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

# Standard ClashKing logo URL
CLASHKING_LOGO_URL = "https://assets.clashk.ing/logos/crown-text-white-bg/BqlEp974170917vB1qK0zunfANJCGi0W031dTksEq7KQ9LoXWMFk0u77unHJa.png"

# Constants
ZERO_PERCENT = "0 (0%)"


async def insert_logo_from_cdn(sheet: Worksheet, image_url: str, anchor_cell="A1", height=80):
    """Insert a logo from CDN URL into an Excel sheet"""
    import asyncio
    import os

    async with aiohttp.ClientSession() as session:
        async with session.get(image_url) as resp:
            if resp.status != 200:
                raise aiohttp.ClientError(f"Failed to download logo from CDN: {resp.status}")
            image_data = await resp.read()

    # Use async-safe temporary file creation
    fd, tmp_img_path = await asyncio.to_thread(
        tempfile.mkstemp,
        suffix=".png",
        prefix="logo_",
        dir=None,
        text=False
    )

    # Write image data and close file descriptor
    await asyncio.to_thread(os.write, fd, image_data)
    await asyncio.to_thread(os.close, fd)

    logo = OpenpyxlImage(tmp_img_path)
    logo.height = height
    logo.width = int(height * 3)
    sheet.add_image(logo, anchor_cell)


async def add_clashking_logo_to_sheet(sheet: Worksheet, anchor_cell="H1", height=50):
    """Add the standard ClashKing logo to a sheet"""
    await insert_logo_from_cdn(
        sheet,
        image_url=CLASHKING_LOGO_URL,
        anchor_cell=anchor_cell,
        height=height
    )


def _get_thin_border() -> Border:
    """Create a thin border for table cells."""
    return Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )


def _get_row_fill(row_num: int, should_highlight: bool) -> PatternFill:
    """Get the fill pattern for a data row."""
    if should_highlight:
        return PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    return CLASHKING_THEME["data_fill_white"] if row_num % 2 == 0 else PatternFill(
        start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"
    )


def _style_header_row(sheet: Worksheet, start_row: int, thin_border: Border) -> None:
    """Apply styling to the header row."""
    for cell in sheet[start_row]:
        cell.value = "" if cell.value is None else str(cell.value)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = CLASHKING_THEME["header_fill_red"]
        cell.font = CLASHKING_THEME["header_font_white"]
        cell.border = thin_border


def _style_data_rows(sheet: Worksheet, start_row: int, end_row: int, thin_border: Border,
                     highlight_rows: Optional[List[int]]) -> None:
    """Apply styling to data rows with alternating colors and highlights."""
    for row_num, row in enumerate(sheet.iter_rows(min_row=start_row + 1, max_row=end_row)):
        actual_row_num = start_row + 1 + row_num
        should_highlight = highlight_rows and actual_row_num in highlight_rows
        row_fill = _get_row_fill(row_num, should_highlight)

        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.fill = row_fill
            cell.border = thin_border


def _auto_adjust_columns(sheet: Worksheet) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in sheet.columns:
        first_real_cell = next((cell for cell in column_cells if isinstance(cell, Cell)), None)
        if not first_real_cell:
            continue
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = first_real_cell.column_letter
        sheet.column_dimensions[col_letter].width = min(length + 2, 50)


def format_table(sheet: Worksheet, start_row: int, end_row: int,
                 highlight_rows: Optional[List[int]] = None):
    """
    Format an Excel table with ClashKing styling

    Args:
        sheet: The worksheet containing the table
        start_row: Row number where the table headers start
        end_row: Row number where the table data ends
        highlight_rows: List of row numbers to highlight with gold color
    """
    thin_border = _get_thin_border()
    _style_header_row(sheet, start_row, thin_border)
    _style_data_rows(sheet, start_row, end_row, thin_border, highlight_rows)
    _auto_adjust_columns(sheet)


def add_title_to_sheet(sheet: Worksheet, title: str, merge_range: str = "A1:P1"):
    """Add a formatted title to a sheet"""
    row = sheet.max_row + 3
    sheet.merge_cells(merge_range.replace("1", str(row)))
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.alignment = Alignment(horizontal="center")
    cell.fill = CLASHKING_THEME["title_fill"]
    cell.font = Font(bold=True, size=14)


def add_filter_summary_to_sheet(sheet: Worksheet, filter_summary: str, merge_range: str = "A1:P1"):
    """Add filter summary under title"""
    row = sheet.max_row + 1
    sheet.merge_cells(merge_range.replace("1", str(row)))
    cell = sheet.cell(row=row, column=1)
    cell.value = filter_summary
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(italic=True, size=10)
    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")


def add_table_headers(sheet: Worksheet, headers: List[str]):
    """Add table headers to a sheet"""
    row = sheet.max_row + 1
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=i)
        cell.value = header
    return row


def add_summary_data(sheet: Worksheet, data: List[tuple], preserve_numeric_types: bool = True):
    """
    Add summary data to a sheet with proper type preservation

    Args:
        sheet: The worksheet
        data: List of (label, value) tuples
        preserve_numeric_types: Whether to keep numeric values as numbers
    """
    for stat_name, stat_value in data:
        if preserve_numeric_types and isinstance(stat_value, (int, float)):
            sheet.append([str(stat_name), stat_value])
        else:
            sheet.append([str(stat_name), str(stat_value)])


def add_insights_section(sheet: Worksheet, insights: List[str], merge_columns: str = "A:I"):
    """Add properly formatted insights as merged cells"""
    for insight in insights:
        row = sheet.max_row + 1
        merge_range = f"{merge_columns[0]}{row}:{merge_columns[-1]}{row}"
        sheet.merge_cells(merge_range)
        cell = sheet.cell(row=row, column=1)
        cell.value = insight
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(italic=True, size=11)
        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue background


def add_section_title(sheet: Worksheet, title: str, merge_columns: str = "A:I"):
    """Add a section title with proper formatting"""
    row = sheet.max_row + 3
    merge_range = f"{merge_columns[0]}{row}:{merge_columns[-1]}{row}"
    sheet.merge_cells(merge_range)
    cell = sheet.cell(row=row, column=1)
    cell.value = title
    cell.alignment = Alignment(horizontal="center")
    cell.fill = CLASHKING_THEME["header_fill_black"]
    cell.font = CLASHKING_THEME["header_font_white"]


def add_empty_row(sheet: Worksheet):
    """Add an empty row for spacing"""
    sheet.append([])


def get_war_league_label(league_id: int) -> str:
    """Get the war league label from league ID.

    Args:
        league_id: The league ID number

    Returns:
        The league name string
    """
    league_mapping = {
        48000000: "Unranked",
        48000001: "Bronze League III",
        48000002: "Bronze League II",
        48000003: "Bronze League I",
        48000004: "Silver League III",
        48000005: "Silver League II",
        48000006: "Silver League I",
        48000007: "Gold League III",
        48000008: "Gold League II",
        48000009: "Gold League I",
        48000010: "Crystal League III",
        48000011: "Crystal League II",
        48000012: "Crystal League I",
        48000013: "Master League III",
        48000014: "Master League II",
        48000015: "Master League I",
        48000016: "Champion League III",
        48000017: "Champion League II",
        48000018: "Champion League I"
    }
    return league_mapping.get(league_id, f"Unknown ({league_id})")


def build_cwl_info_data(cwl_data: dict) -> List[tuple]:
    """Build CWL clan information data for Excel export.

    Args:
        cwl_data: CWL data dictionary

    Returns:
        List of (label, value) tuples
    """
    return [
        ("Clan Name", cwl_data.get('clan_name', 'Unknown')),
        ("Clan Tag", cwl_data.get('clan_tag', 'Unknown')),
        ("Season", cwl_data.get('season', 'Unknown')),
        ("League", get_war_league_label(cwl_data.get('league_id', 0))),
        ("Final Position", cwl_data.get('final_rank', 'Unknown')),
        ("Total Stars", cwl_data.get('total_stars', 0)),
        ("Total Destruction", f"{cwl_data.get('total_destruction', 0):.1f}%"),
        ("Average Stars per Attack", f"{cwl_data.get('average_stars', 0):.2f}"),
    ]


def build_cwl_member_row(member: dict) -> List:
    """Build a member data row for CWL export.

    Args:
        member: Member data dictionary

    Returns:
        List of values for the row
    """
    return [
        member.get('name', 'Unknown'),
        member.get('tag', 'Unknown'),
        member.get('townhall_level', 0),
        member.get('attack_count', 0),
        member.get('total_stars', 0),
        f"{member.get('average_stars', 0):.2f}",
        f"{member.get('total_destruction', 0):.1f}%",
        f"{member.get('average_destruction', 0):.1f}%",
        f"{member.get('performance_score', 0):.2f}"
    ]


def generate_cwl_filename(clan_name: str, clan_tag: str, season: str) -> str:
    """Generate a clean filename for CWL export.

    Args:
        clan_name: Clan name
        clan_tag: Clan tag
        season: Season string

    Returns:
        Clean filename string
    """
    clean_clan_name = "".join(
        c for c in clan_name if c.isalnum() or c in (' ', '-', '_')
    ).rstrip()
    clean_tag = clan_tag.replace('#', '')
    return f"cwl_summary_{clean_clan_name}_{clean_tag}_season_{season}.xlsx"


def generate_player_war_filename(player_name: str, player_tag: str) -> str:
    """Generate a clean filename for player war stats export.

    Args:
        player_name: Player name
        player_tag: Player tag

    Returns:
        Clean filename string
    """
    clean_player_name = "".join(
        c for c in player_name if c.isalnum() or c in (' ', '-', '_')
    ).rstrip()
    clean_tag = player_tag.replace('#', '')
    return f"war_stats_{clean_player_name}_{clean_tag}.xlsx"


def _format_list_or_value(value, prefix: str) -> str:
    """Format a value that can be either a single item or a list."""
    if isinstance(value, list):
        return f"{prefix}: {', '.join(map(str, value))}"
    return f"{prefix}: {value}"


def _add_season_filter(filters: List[str], filter_obj) -> None:
    """Add season filter to filters list if applicable."""
    if filter_obj.season is not None and str(filter_obj.season).strip() and str(filter_obj.season).strip() != 'None':
        filters.append(f"Season {filter_obj.season}")


def _add_war_type_filter(filters: List[str], filter_obj) -> None:
    """Add war type filter to filters list if applicable."""
    if filter_obj.type and filter_obj.type != "all":
        filters.append(_format_list_or_value(filter_obj.type, "War type"))


def _add_th_filters(filters: List[str], filter_obj) -> None:
    """Add townhall filters to filters list if applicable."""
    if filter_obj.own_th is not None:
        filters.append(_format_list_or_value(filter_obj.own_th, "Own TH"))
    if filter_obj.enemy_th is not None:
        filters.append(_format_list_or_value(filter_obj.enemy_th, "Enemy TH"))


def _add_destruction_filters(filters: List[str], filter_obj) -> None:
    """Add destruction filters to filters list if applicable."""
    if filter_obj.min_destruction is not None and filter_obj.min_destruction > 0:
        if filter_obj.max_destruction is not None and filter_obj.max_destruction < 100:
            filters.append(f"Destruction: {filter_obj.min_destruction}%-{filter_obj.max_destruction}%")
        else:
            filters.append(f"Destruction: ≥{filter_obj.min_destruction}%")
    elif filter_obj.max_destruction is not None and filter_obj.max_destruction < 100:
        filters.append(f"Destruction: ≤{filter_obj.max_destruction}%")


def _add_map_position_filters(filters: List[str], filter_obj) -> None:
    """Add map position filters to filters list if applicable."""
    if filter_obj.map_position_min is not None and filter_obj.map_position_min > 1:
        if filter_obj.map_position_max is not None and filter_obj.map_position_max < 50:
            filters.append(f"Map position: {filter_obj.map_position_min}-{filter_obj.map_position_max}")
        else:
            filters.append(f"Map position: ≥{filter_obj.map_position_min}")
    elif filter_obj.map_position_max is not None and filter_obj.map_position_max < 50:
        filters.append(f"Map position: ≤{filter_obj.map_position_max}")


def _add_date_range_filter(filters: List[str], filter_obj) -> None:
    """Add date range filter to filters list if applicable."""
    from datetime import datetime

    if filter_obj.timestamp_start != 0 or filter_obj.timestamp_end != 2527625513:
        start_date = datetime.fromtimestamp(filter_obj.timestamp_start).strftime(
            '%Y-%m-%d') if filter_obj.timestamp_start > 0 else "Beginning"
        end_date = datetime.fromtimestamp(filter_obj.timestamp_end).strftime(
            '%Y-%m-%d') if filter_obj.timestamp_end < 2527625513 else "Present"
        filters.append(f"Date range: {start_date} to {end_date}")


def build_filter_summary(filter_obj) -> str:
    """Build a human-readable summary of applied filters.

    Args:
        filter_obj: Filter object with various filter attributes

    Returns:
        String summary of active filters
    """
    filters = []

    _add_season_filter(filters, filter_obj)
    _add_war_type_filter(filters, filter_obj)
    _add_th_filters(filters, filter_obj)

    if filter_obj.stars is not None:
        filters.append(f"Stars: {', '.join(map(str, filter_obj.stars))}")

    _add_destruction_filters(filters, filter_obj)
    _add_map_position_filters(filters, filter_obj)

    if filter_obj.fresh_only:
        filters.append("Fresh attacks only")

    _add_date_range_filter(filters, filter_obj)

    return "No filters applied" if not filters else "; ".join(filters)


def _parse_war_date(preparation_start_time: str) -> str:
    """Parse war preparation start time to formatted date string."""
    from datetime import datetime

    if not preparation_start_time:
        return ""
    try:
        return datetime.strptime(preparation_start_time, '%Y%m%dT%H%M%S.%fZ').strftime('%Y-%m-%d %H:%M')
    except (ValueError, AttributeError):
        return ""


def _build_attack_dict(attack: dict, war_data: dict, war_date: str) -> dict:
    """Build attack dictionary from attack and war data."""
    return {
        "attacker_name": attack.get("attacker", {}).get("name", ""),
        "attacker_tag": attack.get("attacker", {}).get("tag", ""),
        "attacker_townhall": attack.get("attacker", {}).get("townhallLevel", ""),
        "war_date": war_date,
        "war_type": "CWL" if war_data.get("season") else "Regular",
        "attacker_map_position": attack.get("attacker", {}).get("mapPosition", ""),
        "defender_name": attack.get("defender", {}).get("name", ""),
        "defender_tag": attack.get("defender", {}).get("tag", ""),
        "defender_townhall": attack.get("defender", {}).get("townhallLevel", ""),
        "defender_map_position": attack.get("defender", {}).get("mapPosition", ""),
        "stars": attack.get("stars", 0),
        "destruction_percentage": attack.get("destructionPercentage", 0),
        "order": attack.get("order", ""),
        "fresh_attack": attack.get("freshAttack", attack.get("fresh", False)),
        "war_tag": war_data.get("tag", ""),
        "war_state": war_data.get("state", "")
    }


def extract_war_hits_from_results(player_results: List[dict]) -> List[dict]:
    """Extract individual attack data from player war results.

    Args:
        player_results: List of player war results

    Returns:
        List of attack dictionaries
    """
    war_hits = []
    for player in player_results:
        for war_info in player.get("wars", []):
            war_data = war_info.get("war_data", {})
            member_data = war_info.get("members", [{}])[0]
            war_date = _parse_war_date(war_data.get("preparationStartTime", ""))

            for attack in member_data.get("attacks", []):
                war_hits.append(_build_attack_dict(attack, war_data, war_date))

    return war_hits


class WarStatsExporter:
    """Excel exporter for war statistics"""

    def __init__(self, war_hits: List[Dict[str, Any]], player_tag: str):
        self.war_hits = war_hits
        self.player_tag = player_tag
        self.wb = Workbook()

        # Calculate basic statistics
        self.total_attacks = len(war_hits)
        self.total_stars = sum(hit.get("stars", 0) for hit in war_hits)
        self.total_destruction = sum(hit.get("destruction_percentage", 0) for hit in war_hits)
        self.total_wars = len({hit.get("war_tag", "") for hit in war_hits if hit.get("war_tag")})

        # Star distribution
        self.three_stars = sum(1 for hit in war_hits if hit.get("stars") == 3)
        self.two_stars = sum(1 for hit in war_hits if hit.get("stars") == 2)
        self.one_star = sum(1 for hit in war_hits if hit.get("stars") == 1)
        self.zero_stars = sum(1 for hit in war_hits if hit.get("stars") == 0)

        # TH analysis
        self.th_attacks = {}
        self.opponent_th_attacks = {}
        self.th_stars_breakdown = {}
        self.attack_matchups = {}
        self.defense_matchups = {}

        # Calculate current player TH and matchup data
        self._analyze_th_data()

    @staticmethod
    def _init_matchup_stats() -> Dict[str, Any]:
        """Initialize matchup statistics dictionary."""
        return {0: 0, 1: 0, 2: 0, 3: 0, 'total_destruction': 0, 'attack_count': 0}

    def _update_matchup_stats(self, matchup_dict: Dict, key: Tuple, stars: int, destruction: float) -> None:
        """Update matchup statistics for a given key."""
        if key not in matchup_dict:
            matchup_dict[key] = self._init_matchup_stats()
        matchup_dict[key][stars] += 1
        matchup_dict[key]['total_destruction'] += destruction
        matchup_dict[key]['attack_count'] += 1

    def _analyze_th_data(self):
        """Analyze TH level data and matchups"""
        for hit in self.war_hits:
            own_th = hit.get("attacker_townhall", "Unknown")
            enemy_th = hit.get("defender_townhall", "Unknown")
            stars = hit.get("stars", 0)
            destruction = hit.get("destruction_percentage", 0)

            self.th_attacks[own_th] = self.th_attacks.get(own_th, 0) + 1
            self.opponent_th_attacks[enemy_th] = self.opponent_th_attacks.get(enemy_th, 0) + 1

            # Track stars by TH level
            if own_th not in self.th_stars_breakdown:
                self.th_stars_breakdown[own_th] = {0: 0, 1: 0, 2: 0, 3: 0}
            self.th_stars_breakdown[own_th][stars] = self.th_stars_breakdown[own_th].get(stars, 0) + 1

            # Track attack and defense matchups
            if own_th != "Unknown" and enemy_th != "Unknown":
                self._update_matchup_stats(self.attack_matchups, (own_th, enemy_th), stars, destruction)
                self._update_matchup_stats(self.defense_matchups, (enemy_th, own_th), stars, destruction)

        # Determine current player TH (highest TH level used in attacks)
        self.current_player_th = max((int(th) for th in self.th_attacks.keys() if str(th).isdigit()), default=0)

    async def _create_summary_sheet(self, filter_summary: str) -> None:
        """Create the summary sheet with statistics and analysis"""
        ws_summary = self.wb.active
        ws_summary.title = "Summary"

        # Add logo
        await add_clashking_logo_to_sheet(ws_summary, anchor_cell="D1")

        # Title and filters
        add_title_to_sheet(ws_summary, f"War Statistics Summary for {self.player_tag}", "A1:B1")
        add_filter_summary_to_sheet(ws_summary, filter_summary, "A1:B1")

        # Basic statistics
        summary_data = self._get_summary_data()
        ws_summary.append(["Statistic", "Value"])
        start_summary_row = ws_summary.max_row
        add_summary_data(ws_summary, summary_data, preserve_numeric_types=True)
        end_summary_row = ws_summary.max_row
        format_table(ws_summary, start_summary_row, end_summary_row)

        # TH Attack Performance Table
        if self.th_stars_breakdown:
            self._add_th_performance_table(ws_summary)

    def _get_summary_data(self) -> List[Tuple[str, Any]]:
        """Get summary statistics data"""
        # War type breakdown
        cwl_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") == "CWL")
        regular_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") == "Regular")
        friendly_war_attacks = sum(1 for hit in self.war_hits if hit.get("war_type") not in ["CWL", "Regular"])

        # Fresh vs cleanup attacks
        fresh_attacks = sum(1 for hit in self.war_hits if hit.get("fresh_attack", False))
        cleanup_attacks = self.total_attacks - fresh_attacks

        # Destruction analysis
        high_destruction = sum(1 for hit in self.war_hits if hit.get("destruction_percentage", 0) >= 90)
        medium_destruction = sum(1 for hit in self.war_hits if 50 <= hit.get("destruction_percentage", 0) < 90)
        low_destruction = sum(1 for hit in self.war_hits if hit.get("destruction_percentage", 0) < 50)

        # Most common and targeted TH
        most_common_th = max(self.th_attacks.items(), key=lambda x: x[1]) if self.th_attacks else ("Unknown", 0)
        most_targeted_th = max(self.opponent_th_attacks.items(), key=lambda x: x[1]) if self.opponent_th_attacks else (
            "Unknown", 0)

        # Performance rates
        perfect_rate = round(self.three_stars / self.total_attacks * 100, 2) if self.total_attacks > 0 else 0
        fail_rate = round(self.zero_stars / self.total_attacks * 100, 2) if self.total_attacks > 0 else 0

        return [
            # Top Level Stats
            ["Total Wars", self.total_wars],
            ["Total Attacks", self.total_attacks],
            ["Total Stars", self.total_stars],
            ["", ""],  # Separator

            # War Stats
            ["CWL Attacks", self._format_percentage(cwl_attacks)],
            ["Regular War Attacks", self._format_percentage(regular_attacks)],
            ["Friendly War Attacks", self._format_percentage(friendly_war_attacks)],
            ["", ""],  # Separator

            # Average Stats
            ["Average Stars", round(self.total_stars / self.total_attacks, 2) if self.total_attacks > 0 else 0],
            ["Average Destruction",
             f"{round(self.total_destruction / self.total_attacks, 2)}%" if self.total_attacks > 0 else "0%"],
            ["", ""],  # Separator

            # Attack Distribution
            ["3 Star Attacks", f"{self.three_stars} ({perfect_rate}%)"],
            ["2 Star Attacks", self._format_percentage(self.two_stars)],
            ["1 Star Attacks", self._format_percentage(self.one_star)],
            ["0 Star Attacks", f"{self.zero_stars} ({fail_rate}%)"],
            ["", ""],  # Separator

            # TH Analysis
            ["Most Used TH", f"TH{most_common_th[0]} ({most_common_th[1]} attacks)"],
            ["Most Targeted TH", f"TH{most_targeted_th[0]} ({most_targeted_th[1]} attacks)"],
            ["", ""],  # Separator

            # Destruction Analysis
            ["High Destruction (≥90%)", self._format_percentage(high_destruction)],
            ["Medium Destruction (50-89%)", self._format_percentage(medium_destruction)],
            ["Low Destruction (<50%)", self._format_percentage(low_destruction)],
            ["", ""],  # Separator

            # Attack Type
            ["Fresh Attacks", self._format_percentage(fresh_attacks)],
            ["Cleanup Attacks", self._format_percentage(cleanup_attacks)]
        ]

    @staticmethod
    def _format_star_percentage(count: int, total: int) -> str:
        """Format star count with percentage."""
        if total == 0:
            return ZERO_PERCENT
        return f"{count} ({round(count / total * 100, 1)}%)"

    def _format_percentage(self, count: int, decimals: int = 2) -> str:
        """Format count with percentage based on total attacks."""
        if self.total_attacks == 0:
            return ZERO_PERCENT
        percentage = round(count / self.total_attacks * 100, decimals)
        return f"{count} ({percentage}%)"

    def _add_th_performance_table(self, ws_summary):
        """Add TH Attack Performance Table"""
        add_section_title(ws_summary, "Attack Performance by Town Hall Level", "A:F")

        # Table headers
        headers = ["TH Level", "Total Attacks", "3⭐", "2⭐", "1⭐", "0⭐"]
        th_table_start = add_table_headers(ws_summary, headers)

        # Track rows that should be highlighted (current player TH)
        current_th_rows = []

        # Sort TH levels numerically
        for th in sorted(self.th_stars_breakdown.keys(), key=lambda x: int(x) if str(x).isdigit() else 999):
            if th == "Unknown":
                continue
            stars_data = self.th_stars_breakdown[th]
            total_th_attacks = sum(stars_data.values())
            if total_th_attacks > 0:
                ws_summary.append([
                    f"TH{th}",
                    total_th_attacks,
                    f"{stars_data[3]} ({round(stars_data[3] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[2]} ({round(stars_data[2] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[1]} ({round(stars_data[1] / total_th_attacks * 100, 1)}%)",
                    f"{stars_data[0]} ({round(stars_data[0] / total_th_attacks * 100, 1)}%)"
                ])

                # Mark current player TH row for highlighting
                if int(th) == self.current_player_th:
                    current_th_rows.append(ws_summary.max_row)

        th_table_end = ws_summary.max_row
        format_table(ws_summary, th_table_start, th_table_end, highlight_rows=current_th_rows)

        # Attack Performance Matrix
        if self.attack_matchups:
            self._add_attack_performance_matrix(ws_summary)

        # Defense Performance Matrix
        if self.defense_matchups:
            self._add_defense_performance_matrix(ws_summary)

    @staticmethod
    def _calculate_matchup_stats(stars_data: dict, total_attacks: int) -> Tuple[float, float]:
        """Calculate average stars and destruction for a matchup."""
        avg_stars = (stars_data[3] * 3 + stars_data[2] * 2 + stars_data[1] * 1) / total_attacks
        avg_destruction = stars_data.get('total_destruction', 0) / total_attacks if total_attacks > 0 else 0
        return avg_stars, avg_destruction

    def _build_matchup_row(self, matchup_label: str, stars_data: dict, total_attacks: int,
                           extra_stat: str) -> List:
        """Build a row for matchup matrix table."""
        return [
            matchup_label,
            total_attacks,
            self._format_star_percentage(stars_data[3], total_attacks),
            self._format_star_percentage(stars_data[2], total_attacks),
            self._format_star_percentage(stars_data[1], total_attacks),
            self._format_star_percentage(stars_data[0], total_attacks),
            *extra_stat
        ]

    def _add_attack_performance_matrix(self, ws_summary):
        """Add Attack Performance Matrix"""
        add_section_title(ws_summary, "Attack Performance Matrix")

        headers = ["Matchup", "Attacks", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Efficiency"]
        attack_table_start = add_table_headers(ws_summary, headers)
        current_th_attack_rows = []

        for (your_th, enemy_th) in sorted(self.attack_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = self.attack_matchups[(your_th, enemy_th)]
            total_matchup_attacks = stars_data.get('attack_count',
                                                   sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_matchup_attacks > 0:
                avg_stars, avg_destruction = self._calculate_matchup_stats(stars_data, total_matchup_attacks)
                efficiency = f"{(avg_stars / 3.0 * 100):.0f}%"

                row_data = self._build_matchup_row(
                    f"TH{your_th} vs TH{enemy_th}",
                    stars_data,
                    total_matchup_attacks,
                    [f"{avg_stars:.2f}", f"{avg_destruction:.1f}%", efficiency]
                )
                ws_summary.append(row_data)

                if int(your_th) == self.current_player_th:
                    current_th_attack_rows.append(ws_summary.max_row)

        attack_table_end = ws_summary.max_row
        format_table(ws_summary, attack_table_start, attack_table_end,
                     highlight_rows=current_th_attack_rows)

    def _add_defense_performance_matrix(self, ws_summary):
        """Add Defense Performance Matrix"""
        add_section_title(ws_summary, "Defense Performance Matrix")

        headers = ["Matchup", "Defenses", "3⭐", "2⭐", "1⭐", "0⭐", "Avg Stars", "Avg Destruction", "Defense Rate"]
        defense_table_start = add_table_headers(ws_summary, headers)
        current_th_defense_rows = []

        for (enemy_th, your_th) in sorted(self.defense_matchups.keys(), key=lambda x: (
                int(x[0]) if str(x[0]).isdigit() else 999, int(x[1]) if str(x[1]).isdigit() else 999)):
            stars_data = self.defense_matchups[(enemy_th, your_th)]
            total_defense_attacks = stars_data.get('attack_count',
                                                   sum(v for k, v in stars_data.items() if k in [0, 1, 2, 3]))
            if total_defense_attacks > 0:
                avg_stars_given, avg_destruction_given = self._calculate_matchup_stats(stars_data, total_defense_attacks)

                held_completely = stars_data[0]
                partial_holds = stars_data[1] + stars_data[2]
                defense_rate = f"{((held_completely + partial_holds) / total_defense_attacks * 100):.1f}%"

                row_data = self._build_matchup_row(
                    f"TH{enemy_th} vs TH{your_th}",
                    stars_data,
                    total_defense_attacks,
                    [f"{avg_stars_given:.2f}", f"{avg_destruction_given:.1f}%", defense_rate]
                )
                ws_summary.append(row_data)

                if int(your_th) == self.current_player_th:
                    current_th_defense_rows.append(ws_summary.max_row)

        defense_table_end = ws_summary.max_row
        format_table(ws_summary, defense_table_start, defense_table_end,
                     highlight_rows=current_th_defense_rows)

    async def _create_attack_details_sheet(self, filter_summary: str) -> None:
        """Create the Attack Details sheet"""
        ws_attacks = self.wb.create_sheet(title="Attack Details")

        # Add logo
        await add_clashking_logo_to_sheet(ws_attacks)

        # Title and filters
        add_title_to_sheet(ws_attacks, f"Attack Details for {self.player_tag}")
        add_filter_summary_to_sheet(ws_attacks, filter_summary)

        # Headers
        headers = [
            "Player Name", "Player Tag", "TH Level", "War Date", "War Type",
            "Map Position", "Enemy Name", "Enemy Tag", "Enemy TH", "Enemy Map Position",
            "Stars", "Destruction %", "Attack Order", "Fresh Attack", "War Tag", "War State"
        ]
        start_row = add_table_headers(ws_attacks, headers)

        # Add data rows
        for hit in self.war_hits:
            ws_attacks.append([
                str(hit.get("attacker_name", "")),
                str(hit.get("attacker_tag", "")),
                str(hit.get("attacker_townhall", "")),
                str(hit.get("war_date", "")),
                str(hit.get("war_type", "")),
                str(hit.get("attacker_map_position", "")),
                str(hit.get("defender_name", "")),
                str(hit.get("defender_tag", "")),
                str(hit.get("defender_townhall", "")),
                str(hit.get("defender_map_position", "")),
                hit.get("stars", 0) or 0,
                hit.get("destruction_percentage", 0) or 0,
                str(hit.get("order", "")),
                "Yes" if hit.get("fresh_attack", False) else "No",
                str(hit.get("war_tag", "")),
                str(hit.get("war_state", ""))
            ])

        end_row = ws_attacks.max_row
        format_table(ws_attacks, start_row, end_row)

    async def _create_defense_details_sheet(self, filter_summary: str) -> None:
        """Create the Defense Details sheet"""
        ws_defenses = self.wb.create_sheet(title="Defense Details")

        # Add logo
        await add_clashking_logo_to_sheet(ws_defenses)

        # Title and filters
        add_title_to_sheet(ws_defenses, f"Defense Details for {self.player_tag}")
        add_filter_summary_to_sheet(ws_defenses, filter_summary)

        # Headers for Defense Details
        headers = [
            "Defender Name", "Defender Tag", "Defender TH", "War Date", "War Type",
            "Map Position", "Attacker Name", "Attacker Tag", "Attacker TH", "Attacker Map Position",
            "Stars Given", "Destruction Given %", "Attack Order", "Fresh Attack", "War Tag", "War State"
        ]
        start_row = add_table_headers(ws_defenses, headers)

        # Add defense data rows (same data but from defender perspective)
        for hit in self.war_hits:
            ws_defenses.append([
                str(hit.get("defender_name", "")),
                str(hit.get("defender_tag", "")),
                str(hit.get("defender_townhall", "")),
                str(hit.get("war_date", "")),
                str(hit.get("war_type", "")),
                str(hit.get("defender_map_position", "")),
                str(hit.get("attacker_name", "")),
                str(hit.get("attacker_tag", "")),
                str(hit.get("attacker_townhall", "")),
                str(hit.get("attacker_map_position", "")),
                hit.get("stars", 0) or 0,
                hit.get("destruction_percentage", 0) or 0,
                str(hit.get("order", "")),
                "Yes" if hit.get("fresh_attack", False) else "No",
                str(hit.get("war_tag", "")),
                str(hit.get("war_state", ""))
            ])

        end_row = ws_defenses.max_row
        format_table(ws_defenses, start_row, end_row)

    async def create_excel_export(self, filter_summary: str) -> NamedTemporaryFile:
        """
        Create the complete Excel export with all sheets

        Args:
            filter_summary: Summary of applied filters

        Returns:
            NamedTemporaryFile containing the Excel workbook
        """
        # Create all sheets
        await self._create_summary_sheet(filter_summary)
        await self._create_attack_details_sheet(filter_summary)
        await self._create_defense_details_sheet(filter_summary)

        # Save to temporary file
        tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
        self.wb.save(tmp.name)
        tmp.seek(0)

        return tmp


async def export_player_war_stats_to_excel(
        war_hits: List[Dict[str, Any]],
        player_tag: str,
        filter_summary: str
) -> NamedTemporaryFile:
    """
    Export player war statistics to Excel format

    Args:
        war_hits: List of war hit data
        player_tag: Player tag for the export
        filter_summary: Summary of applied filters

    Returns:
        NamedTemporaryFile containing the Excel workbook
    """
    exporter = WarStatsExporter(war_hits, player_tag)
    return await exporter.create_excel_export(filter_summary)